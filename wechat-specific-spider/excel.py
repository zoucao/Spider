import openpyxl
from openpyxl.styles import Font
import time


class excelUtil(object):
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.sheet = self.wb.create_sheet(title="spider", index=0)
        self.index = 2

    def init_sheet(self):
        font = Font(name='宋体', size=10, bold=False, italic=False, vertAlign=None,
                    underline='none', strike=False, color='0000FF')
        self.sheet.cell(1, 1).value = "日期"
        self.sheet.cell(1, 2).value = "公众号名称"
        self.sheet.cell(1, 3).value = "标题"
        self.sheet.cell(1, 4).value = "链接"
        self.sheet.cell(1, 5).value = "阅读量"
        self.sheet.cell(1, 6).value = "类型"
        self.sheet.cell(1, 7).value = "角度"
        self.sheet.cell(1, 8).value = "发文时间"
        for i in [1,8]:
            self.sheet.cell(1,i).font = font

    def add_column(self, date=None, nick_name=None, title=None, link=None, read_num=None, start_time=None):
        if nick_name is None:
            self.index = self.index + 1
        else:
            self.sheet.cell(self.index, 1, date)
            self.sheet.cell(self.index, 2, nick_name)
            self.sheet.cell(self.index, 3, title)
            self.sheet.cell(self.index, 4).value = '=HYPERLINK("%s","%s")' % (link, link)
            self.sheet.cell(self.index, 5, read_num)
            self.sheet.cell(self.index, 8, start_time)
            self.index = self.index + 1

    def save(self):
        self.wb.save(time.strftime("{}.xlsx".format("%Y-%m-%dT%H:%M:%S", time.localtime())))
